import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .data import ReportSpec


HEADER_FILL = PatternFill("solid", fgColor="0F8D86")
TITLE_FILL = PatternFill("solid", fgColor="E1F4F3")
THIN_BORDER = Border(
    left=Side(style="thin", color="CFE6E3"),
    right=Side(style="thin", color="CFE6E3"),
    top=Side(style="thin", color="CFE6E3"),
    bottom=Side(style="thin", color="CFE6E3"),
)


def _safe_sheet_title(title: str, existing: set[str]) -> str:
    base = title[:31] or "Report"
    candidate = base
    counter = 2
    while candidate in existing:
        suffix = f" {counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def _add_chart(worksheet, report: ReportSpec, header_row: int, data_row_count: int):
    if data_row_count == 0 or report.kind not in {"recurring", "cost"}:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.height = 7
    chart.width = 13
    chart.title = report.title
    if report.kind == "recurring":
        data_column = 2
        category_column = 1
    else:
        data_column = 3
        category_column = 1
    data = Reference(
        worksheet,
        min_col=data_column,
        min_row=header_row,
        max_row=header_row + data_row_count,
    )
    categories = Reference(
        worksheet,
        min_col=category_column,
        min_row=header_row + 1,
        max_row=header_row + data_row_count,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    chart.y_axis.title = worksheet.cell(header_row, category_column).value
    worksheet.add_chart(chart, f"{get_column_letter(len(worksheet[header_row]) + 2)}2")


def render_excel(report: ReportSpec) -> HttpResponse:
    workbook = Workbook()
    workbook.remove(workbook.active)
    existing_names: set[str] = set()
    generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M")

    for sheet_spec in report.sheets:
        title = _safe_sheet_title(sheet_spec.name, existing_names)
        existing_names.add(title)
        worksheet = workbook.create_sheet(title)
        worksheet.sheet_view.rightToLeft = True
        worksheet.freeze_panes = "A5"
        column_count = max(len(sheet_spec.headers), 1)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        worksheet["A1"] = report.title
        worksheet["A1"].font = Font(size=18, bold=True, color="17312D")
        worksheet["A1"].fill = TITLE_FILL
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 30

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
        worksheet["A2"] = report.subtitle
        worksheet["A2"].alignment = Alignment(horizontal="center")
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=column_count)
        worksheet["A3"] = f"تاريخ التوليد: {generated_at}"
        worksheet["A3"].alignment = Alignment(horizontal="center")

        header_row = 4
        for column_index, header in enumerate(sheet_spec.headers, start=1):
            cell = worksheet.cell(header_row, column_index, header)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        for row_index, row in enumerate(sheet_spec.rows, start=header_row + 1):
            for column_index, value in enumerate(row, start=1):
                cell = worksheet.cell(row_index, column_index, value)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
                if "تكلفة" in sheet_spec.headers[column_index - 1] and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

        if sheet_spec.rows:
            worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(column_count)}{header_row + len(sheet_spec.rows)}"

        summary_row = header_row + len(sheet_spec.rows) + 2
        for summary in sheet_spec.summaries:
            worksheet.merge_cells(
                start_row=summary_row,
                start_column=1,
                end_row=summary_row,
                end_column=column_count,
            )
            worksheet.cell(summary_row, 1, summary)
            worksheet.cell(summary_row, 1).font = Font(bold=True, color="17312D")
            worksheet.cell(summary_row, 1).fill = TITLE_FILL
            worksheet.cell(summary_row, 1).alignment = Alignment(horizontal="right")
            summary_row += 1

        for column_index, header in enumerate(sheet_spec.headers, start=1):
            values = [str(header)] + [
                "" if row[column_index - 1] is None else str(row[column_index - 1])
                for row in sheet_spec.rows
            ]
            width = min(max(max((len(value) for value in values), default=10) + 3, 12), 30)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        _add_chart(worksheet, report, header_row, len(sheet_spec.rows))

    buffer = io.BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    response = HttpResponse(
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{report.filename}.xlsx"'
    response["Content-Length"] = str(len(payload))
    return response
