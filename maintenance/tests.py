import json
import tempfile
from io import BytesIO
from datetime import timedelta
from unittest.mock import MagicMock, patch

from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image
from rest_framework import status
from rest_framework.test import APITestCase

from .models import (
    AssignmentNotification,
    CompanyProfile,
    EngineerProfile,
    MaintenanceRequest,
    MaintenanceSpecialty,
    PublicContactInquiry,
    PublicEngineer,
    RequestActivity,
    User,
)
from .notifications import send_transactional_email


class MaintenanceAPITestCase(APITestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username="admin",
            email="admin@example.com",
            role=User.Role.ADMIN,
            is_staff=True,
        )
        self.quality_user = User.objects.create_user(
            username="quality",
            email="quality@example.com",
            role=User.Role.QUALITY_CONTROLLER,
        )
        self.company_user = User.objects.create_user(
            username="client",
            email="client@example.com",
            role=User.Role.CLIENT_COMPANY,
        )
        self.other_company_user = User.objects.create_user(
            username="other-client",
            email="other-client@example.com",
            role=User.Role.CLIENT_COMPANY,
        )
        self.engineer_user = User.objects.create_user(
            username="engineer",
            email="engineer@example.com",
            role=User.Role.ENGINEER,
            first_name="Sara",
            last_name="Naser",
        )
        self.network_engineer_user = User.objects.create_user(
            username="network-engineer",
            email="network-engineer@example.com",
            role=User.Role.ENGINEER,
        )

        self.company = CompanyProfile.objects.create(
            user=self.company_user,
            company_name="Al Hadhra Medical",
            commercial_register="CR-100",
            contact_phone="+218 91 000 0000",
            address="Tripoli",
        )
        self.other_company = CompanyProfile.objects.create(
            user=self.other_company_user,
            company_name="Other Company",
            commercial_register="CR-200",
            contact_phone="+218 91 111 1111",
            address="Benghazi",
        )
        self.engineer = EngineerProfile.objects.create(
            user=self.engineer_user,
            employee_id="ENG-001",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            phone="+218 91 222 2222",
            experience_years=7,
        )
        self.network_engineer = EngineerProfile.objects.create(
            user=self.network_engineer_user,
            employee_id="ENG-002",
            department="Networks",
            specialty=MaintenanceSpecialty.NETWORKS,
            phone="+218 91 333 3333",
            experience_years=5,
        )

    def request_payload(self):
        return {
            "issue_type": MaintenanceSpecialty.ELECTRICITY,
            "priority": MaintenanceRequest.Priority.HIGH,
            "location_details": "Main building, generator room",
            "description": "Power fluctuation in backup generator panel.",
            "preferred_date": (timezone.now() + timedelta(days=1)).isoformat(),
            "is_hazardous": True,
        }

    def create_request(self, **overrides):
        data = self.request_payload()
        data.update(overrides)
        return MaintenanceRequest.objects.create(client_company=self.company, **data)

    def authenticate(self, user):
        self.client.force_authenticate(user=user)


class MaintenanceRequestPermissionTests(MaintenanceAPITestCase):
    def test_company_creates_request_for_own_profile_without_spoofing_company_id(self):
        self.authenticate(self.company_user)
        payload = self.request_payload()
        payload["client_company"] = self.other_company.id

        response = self.client.post(reverse("maintenance-request-list"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        created = MaintenanceRequest.objects.get(id=response.data["id"])
        self.assertEqual(created.client_company, self.company)
        self.assertEqual(created.status, MaintenanceRequest.Status.NEW)

    def test_company_only_sees_own_requests(self):
        own_request = self.create_request()
        MaintenanceRequest.objects.create(
            client_company=self.other_company,
            issue_type=MaintenanceSpecialty.NETWORKS,
            priority=MaintenanceRequest.Priority.MEDIUM,
            location_details="Remote office",
            description="Network outage",
            preferred_date=timezone.now() + timedelta(days=2),
        )
        self.authenticate(self.company_user)

        response = self.client.get(reverse("maintenance-request-list"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        returned_ids = {item["id"] for item in response.data}
        self.assertEqual(returned_ids, {own_request.id})


class WorkflowTransitionTests(MaintenanceAPITestCase):
    def test_illegal_direct_close_returns_bad_request(self):
        maintenance_request = self.create_request()
        self.authenticate(self.quality_user)

        response = self.client.patch(
            reverse("maintenance-request-detail", args=[maintenance_request.id]),
            {"status": MaintenanceRequest.Status.CLOSED},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        maintenance_request.refresh_from_db()
        self.assertEqual(maintenance_request.status, MaintenanceRequest.Status.NEW)

    def test_quality_controller_assigns_and_engineer_can_start_or_wait_only(self):
        maintenance_request = self.create_request()
        self.authenticate(self.quality_user)

        review_response = self.client.patch(
            reverse("maintenance-request-detail", args=[maintenance_request.id]),
            {"status": MaintenanceRequest.Status.UNDER_REVIEW},
            format="json",
        )
        assign_response = self.client.patch(
            reverse("maintenance-request-detail", args=[maintenance_request.id]),
            {"status": MaintenanceRequest.Status.ASSIGNED, "assigned_engineer": self.engineer.id},
            format="json",
        )

        self.assertEqual(review_response.status_code, status.HTTP_200_OK)
        self.assertEqual(assign_response.status_code, status.HTTP_200_OK)

        self.authenticate(self.engineer_user)
        start_response = self.client.patch(
            reverse("maintenance-request-detail", args=[maintenance_request.id]),
            {"status": MaintenanceRequest.Status.IN_PROGRESS},
            format="json",
        )
        complete_response = self.client.patch(
            reverse("maintenance-request-detail", args=[maintenance_request.id]),
            {"status": MaintenanceRequest.Status.COMPLETED},
            format="json",
        )

        self.assertEqual(start_response.status_code, status.HTTP_200_OK)
        self.assertEqual(complete_response.status_code, status.HTTP_400_BAD_REQUEST)
        maintenance_request.refresh_from_db()
        self.assertEqual(maintenance_request.status, MaintenanceRequest.Status.IN_PROGRESS)
        self.assertIsNotNone(maintenance_request.in_progress_at)

    def test_assigning_engineer_with_wrong_specialty_is_rejected(self):
        maintenance_request = self.create_request(status=MaintenanceRequest.Status.UNDER_REVIEW)
        self.authenticate(self.quality_user)

        response = self.client.patch(
            reverse("maintenance-request-detail", args=[maintenance_request.id]),
            {"status": MaintenanceRequest.Status.ASSIGNED, "assigned_engineer": self.network_engineer.id},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        maintenance_request.refresh_from_db()
        self.assertIsNone(maintenance_request.assigned_engineer)
        self.assertEqual(maintenance_request.status, MaintenanceRequest.Status.UNDER_REVIEW)


class DashboardStatisticsTests(MaintenanceAPITestCase):
    def test_dashboard_statistics_use_real_aggregations(self):
        completed_request = self.create_request(
            status=MaintenanceRequest.Status.COMPLETED,
            assigned_engineer=self.engineer,
            assigned_at=timezone.now() - timedelta(hours=5),
            in_progress_at=timezone.now() - timedelta(hours=4, minutes=30),
            completed_at=timezone.now() - timedelta(hours=1),
        )
        completed_request.created_at = timezone.now() - timedelta(hours=6)
        completed_request.save(update_fields=["created_at"])
        MaintenanceRequest.objects.create(
            client_company=self.company,
            issue_type=MaintenanceSpecialty.ELECTRICITY,
            priority=MaintenanceRequest.Priority.LOW,
            location_details="Warehouse",
            description="Lighting issue",
            preferred_date=timezone.now() + timedelta(days=1),
            status=MaintenanceRequest.Status.NEW,
        )
        self.authenticate(self.quality_user)

        response = self.client.get(reverse("dashboard-statistics"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["total_open_requests"], 1)
        self.assertEqual(response.data["completion_rate"], 50.0)
        self.assertEqual(response.data["top_recurring_maintenance_issues"][0]["issue_type"], MaintenanceSpecialty.ELECTRICITY)
        self.assertEqual(response.data["fastest_responding_engineer"]["engineer_id"], self.engineer.id)
        self.assertGreater(response.data["average_resolution_seconds"], 0)
        self.assertEqual(completed_request.status, MaintenanceRequest.Status.COMPLETED)


class PublicEndpointTests(MaintenanceAPITestCase):
    @staticmethod
    def engineer_avatar():
        buffer = BytesIO()
        Image.new("RGB", (32, 32), color="#1f86ec").save(buffer, format="PNG")
        return SimpleUploadedFile("engineer.png", buffer.getvalue(), content_type="image/png")

    def test_public_impact_statistics_do_not_require_authentication(self):
        self.create_request()

        response = self.client.get(reverse("public-impact-statistics"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("total_open_requests", response.data)
        self.assertIn("completion_rate", response.data)

    def test_public_capabilities_advertise_engineer_profile_support(self):
        response = self.client.get(reverse("public-capabilities"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["engineer_profile_version"], 3)
        self.assertTrue(response.data["engineer_avatar_webp"])
        self.assertTrue(response.data["engineer_availability"])
        self.assertTrue(response.data["engineer_device_identity"])
        self.assertTrue(response.data["engineer_profile_editing"])

    def test_public_tracking_returns_limited_ticket_data(self):
        maintenance_request = self.create_request()

        response = self.client.get(reverse("public-request-tracking", args=[maintenance_request.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["id"], maintenance_request.id)
        self.assertEqual(response.data["status"], MaintenanceRequest.Status.NEW)
        self.assertNotIn("description", response.data)
        self.assertNotIn("location_details", response.data)

    def test_public_contact_inquiry_is_persisted_without_authentication(self):
        payload = {
            "contact_name": "Hashim Nabih",
            "company_name": "Maintenance Partner Co",
            "email": "contact@example.com",
            "phone": "+218 91 444 4444",
            "message": "We want to contract the smart maintenance platform.",
        }

        response = self.client.post(reverse("public-contact-inquiry"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(PublicContactInquiry.objects.filter(email="contact@example.com").exists())

    def test_public_maintenance_request_creates_company_and_ticket(self):
        payload = {
            "contact_name": "Hashim Nabih",
            "company_name": "New Facility Co",
            "commercial_register": "CR-900",
            "email": "facility@example.com",
            "phone": "+218 91 555 5555",
            "address": "Tripoli central district",
            "issue_type": MaintenanceSpecialty.HVAC,
            "priority": MaintenanceRequest.Priority.MEDIUM,
            "location_details": "Main lobby cooling unit",
            "description": "Cooling is unstable during working hours.",
            "preferred_date": (timezone.now() + timedelta(days=1)).isoformat(),
            "is_hazardous": False,
        }

        response = self.client.post(reverse("public-maintenance-request-create"), payload, format="json")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        maintenance_request = MaintenanceRequest.objects.get(id=response.data["id"])
        self.assertEqual(maintenance_request.status, MaintenanceRequest.Status.NEW)
        self.assertEqual(maintenance_request.client_company.company_name, "New Facility Co")
        self.assertEqual(maintenance_request.client_company.user.role, User.Role.CLIENT_COMPANY)

    def test_public_engineer_registration_stores_complete_profile_and_avatar(self):
        payload = {
            "name": "Mariam Salem",
            "phone": "+218 91 666 6666",
            "email": "mariam@example.com",
            "department": "Facilities",
            "specialty": MaintenanceSpecialty.HVAC,
            "profession": "HVAC Engineer",
            "experience_years": 8,
            "avatar": self.engineer_avatar(),
            "device_id": "0d64a0b2-39b5-4583-8314-abb1d203c79d",
            "device_label": "Chrome on Windows",
        }

        with tempfile.TemporaryDirectory() as media_root:
            with self.settings(MEDIA_ROOT=media_root):
                response = self.client.post(
                    reverse("public-engineer-list-create"),
                    payload,
                    format="multipart",
                )

                self.assertEqual(response.status_code, status.HTTP_201_CREATED)
                engineer = PublicEngineer.objects.get(pk=response.data["id"])
                self.assertEqual(engineer.email, "mariam@example.com")
                self.assertEqual(engineer.department, "Facilities")
                self.assertEqual(engineer.profession, "HVAC Engineer")
                self.assertEqual(engineer.experience_years, 8)
                self.assertTrue(engineer.avatar.name.endswith(".webp"))
                with Image.open(engineer.avatar.path) as saved_avatar:
                    self.assertEqual(saved_avatar.format, "WEBP")
                self.assertIsNotNone(engineer.device_id_hash)
                self.assertNotEqual(
                    engineer.device_id_hash,
                    "0d64a0b2-39b5-4583-8314-abb1d203c79d",
                )
                self.assertEqual(engineer.device_label, "Chrome on Windows")
                self.assertTrue(engineer.is_available)
                self.assertIn("availability_token", response.data)
                self.assertNotIn("device_id", response.data)

                list_response = self.client.get(reverse("public-engineer-list-create"))
                self.assertNotIn("availability_token", list_response.data[0])

    def test_engineer_session_can_be_restored_from_same_device(self):
        device_id = "f98c58a7-026f-4fc1-8118-38890d250946"
        engineer = PublicEngineer.objects.create(
            name="Ayoub Salem",
            phone="+218 91 111 2222",
            email="ayoub@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.NETWORKS,
            profession="Network Engineer",
            experience_years=3,
            device_id_hash=PublicEngineer.hash_device_id(device_id),
            device_label="Safari on iPhone",
        )

        response = self.client.post(
            reverse("public-engineer-device-session"),
            {"device_id": device_id},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["id"], engineer.id)
        self.assertEqual(response.data["name"], "Ayoub Salem")
        self.assertEqual(response.data["email"], "ayoub@example.com")
        self.assertEqual(response.data["availability_token"], str(engineer.availability_token))
        engineer.refresh_from_db()
        self.assertIsNotNone(engineer.device_last_seen_at)

    def test_same_device_cannot_register_two_engineers(self):
        device_id = "b6df91ff-1ceb-4df3-93bd-d00cbcfc6891"
        PublicEngineer.objects.create(
            name="Existing Engineer",
            phone="+218 91 222 3333",
            email="existing-device@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            profession="Electrical Engineer",
            experience_years=5,
            device_id_hash=PublicEngineer.hash_device_id(device_id),
        )

        response = self.client.post(
            reverse("public-engineer-list-create"),
            {
                "name": "Second Engineer",
                "phone": "+218 91 444 5555",
                "email": "second-device@example.com",
                "department": "Operations",
                "specialty": MaintenanceSpecialty.HVAC,
                "profession": "HVAC Engineer",
                "experience_years": 2,
                "device_id": device_id,
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("device_id", response.data)

    def test_dashboard_can_edit_and_delete_public_engineer(self):
        engineer = PublicEngineer.objects.create(
            name="Engineer Before Edit",
            phone="+218 91 333 4444",
            email="before-edit@example.com",
            department="Support",
            specialty=MaintenanceSpecialty.SOFTWARE,
            profession="Software Engineer",
            experience_years=2,
        )
        detail_url = reverse("public-engineer-detail", args=[engineer.id])

        update_response = self.client.patch(
            detail_url,
            {
                "name": "Engineer After Edit",
                "email": "after-edit@example.com",
                "experience_years": 6,
            },
            format="multipart",
        )

        self.assertEqual(update_response.status_code, status.HTTP_200_OK)
        self.assertEqual(update_response.data["name"], "Engineer After Edit")
        self.assertEqual(update_response.data["experience_years"], 6)

        delete_response = self.client.delete(detail_url)
        self.assertEqual(delete_response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(PublicEngineer.objects.filter(pk=engineer.id).exists())

    def test_engineer_can_update_availability_only_with_management_token(self):
        engineer = PublicEngineer.objects.create(
            name="Ali Omar",
            phone="+218 91 777 0000",
            email="ali@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            profession="Electrical Engineer",
            experience_years=4,
        )
        url = reverse("public-engineer-availability", args=[engineer.id])

        denied = self.client.post(
            url,
            {"availability_token": "00000000-0000-0000-0000-000000000000", "is_available": False},
            format="json",
        )
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)

        response = self.client.post(
            url,
            {"availability_token": str(engineer.availability_token), "is_available": False},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        engineer.refresh_from_db()
        self.assertFalse(engineer.is_available)

    def test_unavailable_public_engineer_cannot_be_assigned(self):
        engineer = PublicEngineer.objects.create(
            name="Unavailable Engineer",
            phone="+218 91 700 0000",
            email="unavailable@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            profession="Electrical Engineer",
            experience_years=5,
            is_available=False,
        )
        maintenance_request = self.create_request(status=MaintenanceRequest.Status.UNDER_REVIEW)

        response = self.client.post(
            reverse("public-admin-request-transition", args=[maintenance_request.id]),
            {
                "status": MaintenanceRequest.Status.ASSIGNED,
                "assigned_public_engineer_id": engineer.id,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        maintenance_request.refresh_from_db()
        self.assertIsNone(maintenance_request.assigned_public_engineer)

    @override_settings(
        ASSIGNMENT_EMAIL_PROVIDER="smtp",
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="EngiFlow <notifications@example.com>",
    )
    def test_assigning_public_engineer_sends_and_logs_email(self):
        mail.outbox.clear()
        engineer = PublicEngineer.objects.create(
            name="Assigned Engineer",
            phone="+218 91 711 0000",
            email="assigned@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            profession="Electrical Engineer",
            experience_years=8,
        )
        maintenance_request = self.create_request(status=MaintenanceRequest.Status.UNDER_REVIEW)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("public-admin-request-transition", args=[maintenance_request.id]),
                {
                    "status": MaintenanceRequest.Status.ASSIGNED,
                    "assigned_public_engineer_id": engineer.id,
                },
                format="json",
            )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        notification = AssignmentNotification.objects.get(request=maintenance_request)
        self.assertEqual(notification.status, AssignmentNotification.Status.SENT)
        self.assertEqual(notification.provider, AssignmentNotification.Provider.SMTP)
        self.assertEqual(notification.recipient_email, "assigned@example.com")
        self.assertEqual(notification.attempts, 1)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(str(maintenance_request.id), mail.outbox[0].subject)
        self.assertIn(maintenance_request.client_company.company_name, mail.outbox[0].body)

    @override_settings(
        ASSIGNMENT_EMAIL_PROVIDER="cloudflare",
        CLOUDFLARE_EMAIL_ACCOUNT_ID="account-id",
        CLOUDFLARE_EMAIL_API_TOKEN="api-token",
        CLOUDFLARE_EMAIL_FROM_ADDRESS="notifications@example.com",
        CLOUDFLARE_EMAIL_FROM_NAME="EngiFlow",
        CLOUDFLARE_EMAIL_REPLY_TO="support@example.com",
    )
    def test_cloudflare_assignment_email_uses_rest_api_payload(self):
        engineer = PublicEngineer.objects.create(
            name="Cloudflare Engineer",
            phone="+218 91 722 0000",
            email="cloudflare-engineer@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            profession="Electrical Engineer",
            experience_years=6,
        )
        maintenance_request = self.create_request(status=MaintenanceRequest.Status.UNDER_REVIEW)
        cloudflare_response = MagicMock()
        cloudflare_response.__enter__.return_value.read.return_value = json.dumps(
            {
                "success": True,
                "errors": [],
                "messages": [],
                "result": {
                    "delivered": ["cloudflare-engineer@example.com"],
                    "permanent_bounces": [],
                    "queued": [],
                },
            }
        ).encode("utf-8")

        with patch("maintenance.notifications.urlopen", return_value=cloudflare_response) as mocked_urlopen:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.post(
                    reverse("public-admin-request-transition", args=[maintenance_request.id]),
                    {
                        "status": MaintenanceRequest.Status.ASSIGNED,
                        "assigned_public_engineer_id": engineer.id,
                    },
                    format="json",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        notification = AssignmentNotification.objects.get(request=maintenance_request)
        self.assertEqual(notification.status, AssignmentNotification.Status.SENT)
        outgoing_request = mocked_urlopen.call_args.args[0]
        payload = json.loads(outgoing_request.data.decode("utf-8"))
        self.assertEqual(payload["to"], "cloudflare-engineer@example.com")
        self.assertEqual(
            payload["from"],
            {"address": "notifications@example.com", "name": "EngiFlow"},
        )
        self.assertEqual(payload["reply_to"], "support@example.com")
        self.assertIn("text", payload)
        self.assertIn("html", payload)

    @override_settings(
        ASSIGNMENT_EMAIL_PROVIDER="brevo",
        BREVO_API_KEY="brevo-api-key",
        BREVO_FROM_ADDRESS="notifications@example.com",
        BREVO_FROM_NAME="EngiFlow",
        BREVO_REPLY_TO="support@example.com",
    )
    def test_brevo_transactional_email_uses_api_payload(self):
        brevo_response = MagicMock()
        brevo_response.__enter__.return_value.read.return_value = json.dumps(
            {"messageId": "<message-id@example.com>"}
        ).encode("utf-8")

        with patch("maintenance.notifications.urlopen", return_value=brevo_response) as mocked_urlopen:
            provider, result = send_transactional_email(
                "engineer@example.com",
                "OTP",
                "Your code is 1234",
                "<p>Your code is <strong>1234</strong></p>",
            )

        self.assertEqual(provider, AssignmentNotification.Provider.BREVO)
        self.assertEqual(result["messageId"], "<message-id@example.com>")
        outgoing_request = mocked_urlopen.call_args.args[0]
        self.assertEqual(outgoing_request.full_url, "https://api.brevo.com/v3/smtp/email")
        payload = json.loads(outgoing_request.data.decode("utf-8"))
        self.assertEqual(payload["to"], [{"email": "engineer@example.com"}])
        self.assertEqual(
            payload["sender"],
            {"email": "notifications@example.com", "name": "EngiFlow"},
        )
        self.assertEqual(payload["replyTo"], {"email": "support@example.com"})
        self.assertIn("htmlContent", payload)


@override_settings(
    ASSIGNMENT_EMAIL_PROVIDER="smtp",
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    PORTAL_OTP_EXPOSE_CODE=True,
    PORTAL_OTP_COOLDOWN_SECONDS=0,
)
class PortalWorkflowTests(MaintenanceAPITestCase):
    def register_company(self, email="portal-company@example.com"):
        request_response = self.client.post(
            reverse("company-portal-request-code"),
            {
                "purpose": "REGISTER",
                "email": email,
                "company_name": "Portal Services",
                "contact_name": "Omar Saleh",
                "commercial_register": "CR-PORTAL-1",
                "phone": "+218 91 800 1000",
                "address": "Tripoli, Libya",
            },
            format="json",
        )
        self.assertEqual(request_response.status_code, status.HTTP_201_CREATED)
        verify_response = self.client.post(
            reverse("company-portal-verify"),
            {
                "challenge_id": request_response.data["challenge_id"],
                "code": request_response.data["debug_code"],
            },
            format="json",
        )
        self.assertEqual(verify_response.status_code, status.HTTP_200_OK)
        return verify_response.data["token"], verify_response.data["profile"]

    def engineer_token(self, engineer):
        request_response = self.client.post(
            reverse("engineer-portal-request-code"),
            {"email": engineer.email},
            format="json",
        )
        self.assertEqual(request_response.status_code, status.HTTP_201_CREATED)
        verify_response = self.client.post(
            reverse("engineer-portal-verify"),
            {
                "challenge_id": request_response.data["challenge_id"],
                "code": request_response.data["debug_code"],
            },
            format="json",
        )
        self.assertEqual(verify_response.status_code, status.HTTP_200_OK)
        return verify_response.data["token"]

    def test_company_registers_by_otp_and_keeps_a_portal_session(self):
        token, profile = self.register_company()

        response = self.client.get(
            reverse("company-portal-dashboard"),
            HTTP_X_PORTAL_TOKEN=token,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(profile["company_name"], "Portal Services")
        self.assertEqual(response.data["profile"]["email"], "portal-company@example.com")
        self.assertEqual(response.data["requests"], [])
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("رمز التحقق", mail.outbox[0].subject)

    @override_settings(PORTAL_OTP_TTL_MINUTES=5)
    def test_portal_code_expires_after_five_minutes(self):
        engineer = PublicEngineer.objects.create(
            name="Five Minute Engineer",
            phone="+218 91 800 4600",
            email="five-minutes@example.com",
            department="Networks",
            specialty=MaintenanceSpecialty.NETWORKS,
            profession="Network Engineer",
            experience_years=3,
        )

        response = self.client.post(
            reverse("engineer-portal-request-code"),
            {"email": engineer.email},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["expires_in_seconds"], 300)

    @override_settings(PORTAL_OTP_EXPOSE_CODE=False)
    def test_portal_code_is_not_exposed_when_debug_response_is_disabled(self):
        engineer = PublicEngineer.objects.create(
            name="Private Code Engineer",
            phone="+218 91 800 4500",
            email="private-code@example.com",
            department="Networks",
            specialty=MaintenanceSpecialty.NETWORKS,
            profession="Network Engineer",
            experience_years=3,
        )

        response = self.client.post(
            reverse("engineer-portal-request-code"),
            {"email": engineer.email},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertNotIn("debug_code", response.data)
        self.assertEqual(len(mail.outbox), 1)

    def test_archived_company_is_hidden_and_can_register_again(self):
        delete_response = self.client.delete(
            reverse("public-admin-company-detail", args=[self.company.id])
        )
        self.assertEqual(delete_response.status_code, status.HTTP_204_NO_CONTENT)
        self.company.refresh_from_db()
        self.company_user.refresh_from_db()
        self.assertTrue(self.company.is_archived)
        self.assertFalse(self.company_user.is_active)

        list_response = self.client.get(reverse("public-company-list"))
        returned_ids = {item["id"] for item in list_response.data}
        self.assertNotIn(self.company.id, returned_ids)

        token, profile = self.register_company(self.company_user.email)
        self.company.refresh_from_db()
        self.company_user.refresh_from_db()
        self.assertFalse(self.company.is_archived)
        self.assertTrue(self.company_user.is_active)
        self.assertEqual(profile["id"], self.company.id)
        self.assertTrue(token)

    def test_company_request_auto_assigns_only_an_available_idle_specialist(self):
        busy_engineer = PublicEngineer.objects.create(
            name="Busy Engineer",
            phone="+218 91 800 2000",
            email="busy@example.com",
            department="Networks",
            specialty=MaintenanceSpecialty.NETWORKS,
            profession="Network Engineer",
            experience_years=4,
        )
        idle_engineer = PublicEngineer.objects.create(
            name="Idle Engineer",
            phone="+218 91 800 3000",
            email="idle@example.com",
            department="Networks",
            specialty=MaintenanceSpecialty.NETWORKS,
            profession="Network Engineer",
            experience_years=6,
        )
        MaintenanceRequest.objects.create(
            client_company=self.company,
            issue_type=MaintenanceSpecialty.NETWORKS,
            priority=MaintenanceRequest.Priority.HIGH,
            location_details="Existing network room",
            description="Existing active assignment",
            preferred_date=timezone.now() + timedelta(days=1),
            status=MaintenanceRequest.Status.IN_PROGRESS,
            assigned_public_engineer=busy_engineer,
            assigned_at=timezone.now(),
            in_progress_at=timezone.now(),
        )
        token, _ = self.register_company()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("company-portal-request-create"),
                {
                    "issue_type": MaintenanceSpecialty.NETWORKS,
                    "priority": MaintenanceRequest.Priority.MEDIUM,
                    "location_details": "Second floor network cabinet",
                    "description": "Intermittent connection on the second floor.",
                    "preferred_date": (timezone.now() + timedelta(days=2)).isoformat(),
                    "is_hazardous": False,
                },
                format="json",
                HTTP_X_PORTAL_TOKEN=token,
            )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["status"], MaintenanceRequest.Status.ASSIGNED)
        self.assertEqual(response.data["assigned_public_engineer"], idle_engineer.id)
        self.assertTrue(
            RequestActivity.objects.filter(
                request_id=response.data["id"],
                event_type=RequestActivity.EventType.AUTO_ASSIGNED,
            ).exists()
        )

    def test_engineer_updates_availability_status_and_notes_from_portal(self):
        engineer = PublicEngineer.objects.create(
            name="Portal Engineer",
            phone="+218 91 800 4000",
            email="portal-engineer@example.com",
            department="Operations",
            specialty=MaintenanceSpecialty.ELECTRICITY,
            profession="Electrical Engineer",
            experience_years=7,
        )
        maintenance_request = self.create_request(
            status=MaintenanceRequest.Status.ASSIGNED,
            assigned_public_engineer=engineer,
            assigned_at=timezone.now(),
        )
        token = self.engineer_token(engineer)

        availability_response = self.client.post(
            reverse("engineer-portal-availability"),
            {"is_available": False},
            format="json",
            HTTP_X_PORTAL_TOKEN=token,
        )
        action_response = self.client.post(
            reverse("engineer-portal-request-action", args=[maintenance_request.id]),
            {
                "status": MaintenanceRequest.Status.IN_PROGRESS,
                "note": "Inspected the electrical panel and started diagnostics.",
            },
            format="json",
            HTTP_X_PORTAL_TOKEN=token,
        )

        self.assertEqual(availability_response.status_code, status.HTTP_200_OK)
        self.assertFalse(availability_response.data["is_available"])
        self.assertEqual(action_response.status_code, status.HTTP_200_OK)
        self.assertEqual(action_response.data["status"], MaintenanceRequest.Status.IN_PROGRESS)
        self.assertEqual(len(action_response.data["activities"]), 2)
        maintenance_request.refresh_from_db()
        self.assertIsNotNone(maintenance_request.in_progress_at)


class PublicReportTests(MaintenanceAPITestCase):
    def setUp(self):
        super().setUp()
        self.public_engineer = PublicEngineer.objects.create(
            name="Ahmed Ali",
            phone="+218 91 777 7777",
            specialty=MaintenanceSpecialty.ELECTRICITY,
        )
        self.maintenance_request = self.create_request(
            status=MaintenanceRequest.Status.COMPLETED,
            assigned_public_engineer=self.public_engineer,
            completed_at=timezone.now(),
            cost="1250.50",
        )

    def test_admin_can_set_and_clear_request_cost(self):
        url = reverse("public-admin-request-cost", args=[self.maintenance_request.id])

        response = self.client.post(url, {"cost": "900.25"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["cost"], "900.25")

        response = self.client.post(url, {"cost": None}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNone(response.data["cost"])

    def test_negative_cost_is_rejected(self):
        response = self.client.post(
            reverse("public-admin-request-cost", args=[self.maintenance_request.id]),
            {"cost": "-1"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_all_report_kinds_export_pdf_and_excel(self):
        for kind in ["monthly", "company", "engineer", "recurring", "cost"]:
            with self.subTest(kind=kind, file_format="pdf"):
                response = self.client.get(reverse("public-report", args=[kind]), {"file_format": "pdf"})
                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response["Content-Type"], "application/pdf")
                self.assertTrue(response.content.startswith(b"%PDF"))

            with self.subTest(kind=kind, file_format="xlsx"):
                response = self.client.get(reverse("public-report", args=[kind]), {"file_format": "xlsx"})
                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(
                    response["Content-Type"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                workbook = load_workbook(BytesIO(response.content), read_only=True)
                self.assertGreaterEqual(len(workbook.sheetnames), 1)
                self.assertTrue(workbook[workbook.sheetnames[0]]["A1"].value)

    def test_invalid_report_parameters_return_bad_request(self):
        response = self.client.get(
            reverse("public-report", args=["monthly"]),
            {"file_format": "csv", "month": "13"},
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
